# Bibliotecas
import datetime
import re
import openpyxl
import pymysql

#Conneca ao banco de dados
connectiondb = pymysql.connect(
    host='10.199.102.163',
    user='root',
    passwd='eletro123',
    database='cln'
)

# Trata texto de data
data = datetime.datetime.now()
data_formatada = data.strftime('%Y%m%d')
ano = data.strftime('%Y')
mes = data.strftime('%m')
data_anterior = data - datetime.timedelta(days=1)
text_data_anterior = data_anterior.strftime('%Y%m%d')
text_data_bd = data_anterior.strftime('%Y-%m-%d %H:%M:%S')
dia = data_anterior.strftime('%d')

# Destino de Origem e Saida
#nome_arquivo = 'Console-20230605.log'
nome_arquivo = 'R:/logs/SERVIDOR_CCP_NACALA/'+ano+'/'+mes+'/'+dia+'/'+'Console-'+text_data_anterior+'.log'
nome_arquivo_excel = 'transicao.xlsx'

# Percorre o arquivo de logs e conta eventos
def contar_eventos_transicao(texto):
    eventos_transicao = {}
    regex_evento = re.compile(
        r"(Maquina de Chave (\w+) em Transicao)|(Indicação recebida: Maquina de Chave (\w+) em (Transicao))")

    linhas = texto.split('\n')
    for linha in linhas:
        matches = regex_evento.findall(linha)
        for match in matches:
            maquina_chave = match[1] or match[3]
            if maquina_chave in eventos_transicao:
                eventos_transicao[maquina_chave] += 1
            else:
                eventos_transicao[maquina_chave] = 1

    return eventos_transicao
#--------------------------------------------------------------------------------------------------------------

# Ler os dados de um arquivo caso não haja erro
try:
    with open(nome_arquivo, 'r', encoding='latin-1') as file:
        texto_dados = file.read()
    eventos_transicao = contar_eventos_transicao(texto_dados)

    # Criar um novo arquivo Excel e adicionar os dados
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Eventos Transicao'

    # Escrever os cabeçalhos
    sheet['A1'] = 'chave'
    sheet['B1'] = 'qtd'
    sheet['C1'] = 'data'

    # Escrever os dados no excel
    row = 2
    for maquina_chave, total_eventos in eventos_transicao.items():
        try:
            sheet.cell(row=row, column=1).value = maquina_chave
            sheet.cell(row=row, column=2).value = total_eventos
            sheet.cell(row=row, column=3).value = text_data_bd
            row += 1

            # Insere dados no banco de dados
            maquina_chave = maquina_chave\
                .replace("W16A", "6")\
                .replace("W16B", "7")\
                .replace("W10A", "8")\
                .replace("W20", "10")\
                .replace("W14A", "11")\
                .replace("W22", "12")\
                .replace("W21", "13")\
                .replace("W10B", "14")\
                .replace("W30A", "15")\
                .replace("W24", "16")\
                .replace("W23", "17")\
                .replace("W26", "18")\
                .replace("W14B", "19")\
                .replace("W25", "20")\
                .replace("W12B", "21")\
                .replace("W28", "22")\
                .replace("W27", "23")\
                .replace("W12A", "24")\
                .replace("W19B", "25")\
                .replace("W19A", "26")\
                .replace("W17B", "28")\
                .replace("W17A", "29")\
                .replace("W11B", "30")\
                .replace("W11A", "31")\
                .replace("W18", "34")\
                .replace("W15", "35")\
                .replace("W13A", "36")\
                .replace("W13B", "37")\
                .replace("W29", "38")\
                .replace("W30B", "39")\
                .replace("W32B", "40")\
                .replace("W31", "41")\
                .replace("W32A", "42")\
                .replace("W33", "44")\
                .replace("W34", "45")\
                .replace("W8", "32")\
                .replace("W9A", "27")\
                .replace("W7", "33")\
                .replace("W2", "5")\
                .replace("W6", "43")\
                .replace("W9B", "9")\
                .replace("W5", "1")\
                .replace("W3", "2")\
                .replace("W4", "3")\
                .replace("W1", "4")
            inputdb = connectiondb.cursor()
            SQL_command = 'INSERT INTO mv_movimentacao_chave (id_chave,qtd, data) VALUES (%s,%s,%s)'
            data = (int(maquina_chave), str(total_eventos), str(text_data_bd))
            inputdb.execute(SQL_command, data)
            connectiondb.commit()
            print(data)
        except Exception as e:
            print("Ocorreu um erro ao inserir os dados no banco de dados:", str(e))
            continue

    # Salvar a planilha Excel
    workbook.save(nome_arquivo_excel)

except Exception as e:
    print("Ocorreu um erro:", str(e))